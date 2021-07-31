#pip install requests
import requests  
#pip install openpyxl   
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
#pip install beautifulsoup4
from bs4 import BeautifulSoup

# we make vareble for storeg the elements
job = []
client = []
time1 = []
offer = []
desc = []

def get_info():
    
    # this loop for the repertnal all above to athre page
    for h in range(1,16):

        # we include this library for get the link
        result = requests.get(f"https://mostaql.com/projects?page={h}&category=development&budget_max=10000&sort=latest")
        
        # dim vareble for take content
        src = result.content

        # here we include library to edit in html
        suop = BeautifulSoup(src ,"lxml")

        # here we go to web said and we get the tags we need it
        job_titles = suop.find_all("span" , { "class":"text-zeta12 text-meta"})
        client_name = suop.find_all("a" , {"class":"text-muted"})
        time = suop.find_all("time")
        offers = suop.find_all("li" , {"class":"text-muted"})
        description = suop.find_all("a" , {"class":"details-url"})

        # this for to extract the text 
        for i in range(len(job_titles)):
            job.append(job_titles[i].text)
            client.append(client_name[i].text)
            time1.append(time[i].text)
            desc.append(description[i].text)
        for j in range(len(offers)):
            offer.append(offers[j].text)
        print(f"the {h} page is complet")
get_info()
# here we biging use the library openpyxl for edit in our file

# we write the titles for column
ws.append(["عنوان المشروع","اسم العميل", "مدة الطلب", "عدد العروض", "الوصف"])

#this loop to wite in cilumn because this library it write onle row or one element
for i in range(len(job)):
    ws["A"+str(i+2)] = job[i]
    ws["B"+str(i+2)] = client[i]
    ws["C"+str(i+2)] = time1[i]
    ws["E"+str(i+2)] = desc[i]

#this codes with loop because the tag in html is content 2 element then we need only offer
g = 1  
for i in range(len(offer)):
    if ((i%2) != 0):
       g += 1
       ws["D"+str(g)] = offer[i]  

# to save work in my file
wb.save('jobs.xlsx')
